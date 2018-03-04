import openpyxl

wb = openpyxl.load_workbook('DB_Sample.xlsx')
print(wb.get_sheet_names())

ws = wb.get_sheet_by_name('Sheet1')
print(ws.max_column)
print(ws.min_column)
print(ws.max_row)
print(ws.min_row)

#for row in ws.rows:
#    for cell in row:
#        print(cell.value)

#for column in ws.columns:
#    for cell in column:
#        print(cell.value)

